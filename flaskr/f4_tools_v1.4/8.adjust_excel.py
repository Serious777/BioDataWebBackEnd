# coding:utf-8
# @Time : 2021/10/6 13:29 
# @Author : cewinhot 
# @Version：V 0.1
# @File : adjust_excel.py
'''
调整excel格式
指定excel所在文件夹/excel文件
如果未指定则调整当前目录下所有excel文件
'''
import openpyxl
import argparse
import os
from openpyxl.styles import Font, Alignment


def adjust(path, head):
    wb = openpyxl.load_workbook(path)
    for ws in wb:
        row = ws.max_row
        col = ws.max_column
        line = [1 for _ in range(col)]
        for i in range(row):
            for j in range(col):
                try:
                    length = len(ws.cell(row=i + 1, column=j + 1).value)
                except:
                    length = 0
                if length > line[j]:
                    line[j] = length
                ws.cell(row=i + 1, column=j + 1).font = Font(name='Times New Roman', size=12)
                ws.cell(row=i + 1, column=j + 1).alignment = Alignment(vertical='center', horizontal='center')
        for j in range(col):
            ws.column_dimensions[chr(j + 65)].width = line[j] + 3
        if head:
            for j in range(col):
                ws.cell(row=1, column=j + 1).font = Font(name='Times New Roman', size=13, bold=True)
    wb.save(path)


dirs = False
parser = argparse.ArgumentParser()
parser.add_argument('-p', '--path', help='specify the excel file or directory', type=str)
parser.add_argument('--head', help='if need header', action='store_true')
args = parser.parse_args()
if args.path:
    path = args.path
    if not path.endswith('.xlsx'):
        dirs = path
else:
    dirs = os.getcwd()
head = True if args.head else False

if dirs:
    xlsx = [file for file in os.listdir(dirs) if file.endswith('.xlsx')]
    for path in xlsx:
        adjust(path, head)
else:
    adjust(path, head)
