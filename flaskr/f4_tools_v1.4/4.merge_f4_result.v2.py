# coding:utf-8
# @Time : 2021/9/13 1:02
# @Author : cewinhot
# @Version：2.1
# @File : merge_f4_result.py

import os
import openpyxl
import argparse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def detect_len(line, length, len1):
    for i in range(len1):
        t = len(line[i + 1])
        if t > length[i]:
            length[i] = t


def fill_color(value):
    if value >= 3:
        return col[0]
    elif value >= 2:
        return col[1]
    elif value <= -3:
        return col[3]
    elif value <= -2:
        return col[2]
    else:
        return ""


def cell_search_text(cell, text):
    if text in cell.value:
        return True
    else:
        return False


def cell_pattern(cell, target=None, color=None, bold=False, size=None, font=None, italic=None, horizontal=None,
                 vertical=None):
    '''不定义target则直接定义表格样式
       若定义target则判断cell中是否包含target有则自定义表格样式'''
    if target == None or cell_search_text(cell, target):
        if color:
            cell.fill = PatternFill('solid', fgColor=color)
        if bold:
            cell.font = Font(bold=bold)
        if size:
            cell.font = Font(size=size)
        if font:
            cell.font = Font(name=font)
        if italic:
            cell.font = Font(italic=italic)
        if horizontal:
            cell.alignment = Alignment(horizontal=horizontal)
        if vertical:
            cell.alignment = Alignment(vertical=vertical)
        return


# Default setting
if __name__ == '__main__':
    path = r'./'
    suffix = '.result'
    output = 'result.xlsx'
    head = ["pop1", "pop2", "pop3", "pop4", "f4", "std.err", "Z", "ABAB", "ABBA", "SNPs"]
    col = ["82b1ff", "81d4fa", "ffcc80", "ff8a80"]
    border = Border(bottom=Side(border_style='medium', color='000000'))  # 自定义边框格式
    # top=Side(border_style='thick', color='000000'))
    # right=Side(border_style='dashed', color='000000'))
    # left=Side(border_style='thin', color='000000'))
    Usage = '''
    '''
    parser = argparse.ArgumentParser(description=Usage)
    parser.add_argument('-p', '--path', help='specify directory', type=str)
    parser.add_argument('-s', '--suffix', help='specify input  SUFFIX', type=str)
    parser.add_argument('-o', '--output', help='specify output PREFIX', type=str)
    args = parser.parse_args()
    if args.path:
        path = args.path
    if args.suffix:
        suffix = args.suffix
        if suffix[0] != '.':
            suffix = '.' + suffix
    if args.output:
        output = args.output
        if ".xlsx" not in output:
            output += ".xlsx"
    files = [x for x in os.listdir(path) if x.endswith(suffix)]
    # 尝试按数字排序，如果失败则按字母排序
    try:
        files.sort(key=lambda x: int(x.split(suffix)[0]))
    except ValueError:
        print("Warning: Could not sort files numerically, using alphabetical sort")
        files.sort()
    wb = openpyxl.Workbook()

    if os.path.exists(os.path.join(path, "summ.table")):
        f = os.path.join(path, "summ.table")
        with open(f) as text:
            summ = ["#", "Population", "Min Z-score", "Max Z-score", "Min Fst", "Max Fst", "Abs(Min) + Abs(Max)"]
            sheet = wb.create_sheet("summary")
            sheet.append(summ)
            [sheet.append(line.split(sep="\t")) for line in text]

    for file in files:  # 对每个result循环
        f = os.path.join(path, file)
        with open(f) as text:
            sheet = wb.create_sheet(file.split(suffix)[0])
            # 设置表头字体,字号,加粗,居中
            len1 = len(head)
            sheet.append(head)
            for i in range(len1):
                sheet.cell(row=1, column=i + 1).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=1, column=i + 1).font = Font(name='Times New Roman', size=13, bold=True)
            # 赋值, 居中, 设置边框
            t = 2
            length = [0 for _ in range(len1)]  # 保存列宽的列表
            len1 = len(length)
            for i in text:
                i = i.split()
                detect_len(i, length, len1)  # 计算列宽
                sheet.append(i[1:])  # 赋值
                for j in range(len1):  # 居中
                    sheet.cell(row=t, column=j + 1).alignment = Alignment(horizontal='center', vertical='center')
                t += 1
            # for k in range(1, t + 1):  # 设置边框
                # if k % 3 == 1:
                #     for j in range(len1):
                #         sheet.cell(row=k, column=j + 1).border = border
            # 设置列宽
            for i in range(len1):
                sheet.column_dimensions[chr(i + 65)].width = length[i] + 3
            # 上色
            h = len(sheet['G'])
            for i in range(2, h + 1):
                cell = sheet.cell(row=i, column=7)
                m = fill_color(float(cell.value))  # 定义表格颜色
                if m != '':
                    cell_pattern(cell, color=m)
            # 自定义目标群体
            # for i in [2, ]:  # 目标列
            #     for j in range(2, h + 1):  # 目标行
            #         cell = sheet.cell(row=j, column=i)
            #         cell_pattern(cell, target='Yamnaya', color='d7e3bc', bold=False, italic=False, horizontal=None,
            #                      vertical=None)

    del wb['Sheet']  # 删除原始表格
    wb.save(os.path.join(path, output))
    
    # 生成plot.txt文件
    try:
        print("Generating plot.txt...")
        # 首先检查输入文件
        if not files:
            print("No input files found")
            raise Exception("No input files found")
        
        # 读取summ.txt而不是原始result文件
        if not os.path.exists("../summ.txt"):
            print("summ.txt not found")
            raise Exception("summ.txt not found")
        
        data_written = False  # 标记是否写入了任何数据
        with open("plot.txt", "w") as plot_file:
            with open("../summ.txt") as text:
                for line in text:
                    if line.startswith("result:"):
                        fields = line.strip().split()
                        if len(fields) >= 8:  # 确保有足够的字段
                            try:
                                # 格式：pop Fst z_score order
                                pop = f"{fields[2]}-{fields[3]}"  # pop1-pop2
                                fst = fields[5]  # f4值
                                z_score = fields[7]  # Z值
                                plot_file.write(f"{pop}\t{fst}\t{z_score}\t1\n")
                                data_written = True
                            except Exception as e:
                                print(f"Error processing line: {line}")
                                print(f"Fields: {fields}")
                                print(f"Error: {str(e)}")
                                continue
        
        if not data_written:
            raise Exception("No valid data found in summ.txt")
        
        print("Successfully generated plot.txt")
        # 验证生成的文件
        if os.path.exists("plot.txt"):
            with open("plot.txt") as f:
                content = f.read()
                if not content.strip():
                    raise Exception("plot.txt is empty")
                print("plot.txt content:")
                print(content[:200] + "..." if len(content) > 200 else content)
        else:
            print("Failed to generate plot.txt")
            raise Exception("Failed to generate plot.txt")
    except Exception as e:
        print(f"Error generating plot.txt: {str(e)}")
        print(f"Current directory: {os.getcwd()}")
        print("Directory contents:")
        print(os.listdir("."))
        print("Content of summ.txt:")
        try:
            with open("../summ.txt") as f:
                print(f.read())
        except Exception as read_error:
            print(f"Could not read summ.txt: {str(read_error)}")
        raise

