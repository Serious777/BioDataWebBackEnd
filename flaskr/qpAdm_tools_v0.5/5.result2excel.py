#!/usr/bin/env python3
"""
将qpAdm结果转换为Excel格式
"""

import os
import sys
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

def parse_result_line(line):
    """解析结果行，返回组合的结果数据"""
    # 提取目标群体和源群体
    pops = line.split('best coefficients:')[0].strip().split()
    target = pops[0]
    sources = pops[1:]
    
    # 提取系数
    coeffs = re.findall(r'best coefficients:\s*([-\d\.\s]+)', line)
    if coeffs:
        coeffs = [float(x) for x in coeffs[0].strip().split()]
    else:
        return None
        
    # 提取标准误差
    stds = re.findall(r'std\. errors:\s*([-\d\.\s]+)', line)
    if stds:
        stds = [float(x) for x in stds[0].strip().split()]
    else:
        return None
        
    # 提取tail值
    tail = re.findall(r'tail:\s*([-\d\.]+)', line)
    if not tail:
        return None
    tail = float(tail[0])
    
    return {
        'target': target,
        'sources': sources,
        'coeffs': coeffs,
        'stds': stds,
        'tail': tail
    }

def process_result():
    # 检查输入文件
    if not os.path.exists("4.all_result.txt"):
        print("错误: 未找到输入文件 4.all_result.txt")
        return False
        
    # 读取并检查输入文件内容
    with open("4.all_result.txt") as f:
        lines = f.readlines()
    
    if not lines:
        print("错误: 输入文件为空")
        return False
        
    print(f"读取到 {len(lines)} 行结果")
    
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "all"
    ws_suc = wb.create_sheet("success")
    
    # 添加表头
    HEAD = ["target", "source1", "source2", "source3", "ratio1", "ratio2", "ratio3", 
            "std.error1", "std.error2", "std.error3", "tail", 
            "Jackknife1", "Jackknife2", "Jackknife3"]
    ws_all.append(HEAD)
    ws_suc.append(HEAD)
    
    # 处理每行数据
    row_count = 0
    for line in lines:
        result = parse_result_line(line)
        if result:
            # 准备Excel行数据
            row_data = [result['target']]  # target
            
            # sources (最多3个)
            for i in range(3):
                row_data.append(result['sources'][i] if i < len(result['sources']) else '')
                
            # coefficients (最多3个)
            for i in range(3):
                row_data.append(result['coeffs'][i] if i < len(result['coeffs']) else '')
                
            # standard errors (最多3个)
            for i in range(3):
                row_data.append(result['stds'][i] if i < len(result['stds']) else '')
                
            # tail value
            row_data.append(result['tail'])
            
            # Jackknife (暂时留空)
            row_data.extend(['', '', ''])
            
            # 写入all表
            ws_all.append(row_data)
            row_count += 1
            
            # 如果tail > 0.05，写入success表
            if result['tail'] > 0.05:
                ws_suc.append(row_data)
    
    print(f"处理了 {row_count} 行数据")
    
    # 保存Excel文件
    wb.save("qpadm_result.xlsx")
    return True

if __name__ == "__main__":
    if process_result():
        print("成功生成Excel文件")
    else:
        print("生成Excel文件失败")
        sys.exit(1)
