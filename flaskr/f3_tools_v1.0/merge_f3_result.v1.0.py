import os
import openpyxl
import argparse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import logging

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def detect_len(line, length, len1):
    for i in range(len1):
        t = len(line[i + 1])
        if t > length[i]:
            length[i] = t

def fill_color(value):
    try:
        value = float(value)
        if value >= 3:
            return col[0]
        elif value >= 2:
            return col[1]
        elif value <= -3:
            return col[3]
        elif value <= -2:
            return col[2]
        else:
            return ""
    except ValueError:
        logger.error(f"Invalid value for coloring: {value}")
        return ""

def process_result_file(file_path, wb, head):
    try:
        if not os.path.exists(file_path):
            logger.error(f"Result file not found: {file_path}")
            return False

        # 获取当前目录名作为sheet名
        #sheet_name = os.path.basename(os.getcwd())
        sheet_name = "f3_summ_result"
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)

        # 设置表头
        len1 = len(head)
        sheet.append(head)
        for i in range(len1):
            cell = sheet.cell(row=1, column=i + 1)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Times New Roman', size=13, bold=True)

        # 处理数据
        row = 2
        length = [0] * len1
        with open(file_path, 'r') as text:
            for line in text:
                if 'result:' not in line:  # 只处理包含结果的行
                    continue
                try:
                    data = line.split()
                    if len(data) < len1 + 1:
                        continue
                    
                    detect_len(data, length, len1)
                    sheet.append(data[1:])
                    
                    for j in range(len1):
                        cell = sheet.cell(row=row, column=j + 1)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    row += 1
                except Exception as e:
                    logger.error(f"Error processing line: {line}, Error: {str(e)}")
                    continue

        # 设置列宽
        for i in range(len1):
            sheet.column_dimensions[chr(i + 65)].width = length[i] + 3

        # 设置颜色
        for i in range(2, row):
            try:
                cell = sheet.cell(row=i, column=6)  # Z-score 列
                color = fill_color(cell.value)
                if color:
                    cell.fill = PatternFill('solid', fgColor=color)
            except Exception as e:
                logger.error(f"Error applying color to row {i}: {str(e)}")

        return True

    except Exception as e:
        logger.error(f"Error processing file {file_path}: {str(e)}")
        return False

if __name__ == '__main__':
    try:
        # 设置基本参数
        head = ["Source1", "Source2", "Target", "f3", "std.err", "Z", "SNPs"]
        col = ["82b1ff", "81d4fa", "ffcc80", "ff8a80"]
        
        # 创建工作簿并处理结果
        wb = openpyxl.Workbook()
        
        # 确保至少有一个工作表可见
        if 'Sheet' in wb.sheetnames:
            sheet = wb['Sheet']
        else:
            sheet = wb.create_sheet('Sheet')

        # 处理当前目录下的summ.result文件
        if process_result_file('./summ.result', wb, head):
            # 如果处理成功，删除默认的Sheet（如果存在且不是唯一的工作表）
            if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb['Sheet']
            
            # 保存结果
            wb.save('result.xlsx')
            logger.info("Results saved successfully")
        else:
            logger.error("Failed to process results")
            exit(1)

    except Exception as e:
        logger.error(f"Program error: {str(e)}")
        exit(1)

